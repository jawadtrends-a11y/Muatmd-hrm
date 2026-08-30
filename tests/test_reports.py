"""حرّاس التقارير وتصديرها (ق-40، ق-43)."""
import json
from datetime import date, timedelta
from decimal import Decimal as D

import pytest
from django.contrib.auth.models import User
from django.test import Client

from apps.accounts.models import Account, Company
from apps.accounts.models_access import AccountMembership, Role, RoleAssignment
from apps.accounts.services.provisioning import provision_account
from apps.core.access.catalog import Scope
from apps.core.reports import REGISTRY, catalog, get_report, load_reports
from apps.core.reports.base import ReportError
from apps.core.tenancy.context import account_scope
from apps.employees.services.assets import add_document, assign_asset
from apps.employees.services.hiring import create_employment, create_person
from apps.payroll.models import PayComponent, PayrollSettings
from apps.payroll.services.gosi_seed import sync_gosi_rates

IBAN = "SA6080000247608010330101"


@pytest.fixture
def env(db):
    sync_gosi_rates()
    load_reports()
    r = provision_account(slug="rpt-test", display_name_ar="حساب",
                          company_name_ar="شركة التقارير", is_sandbox=True)
    with account_scope(r.account_id):
        acc = Account.objects.get(id=r.account_id)
        comp = Company.objects.get(id=r.company_id)
        comps = {c.code: c for c in PayComponent.objects.filter(company=comp)}
        st = PayrollSettings.objects.get(company=comp)
        st.eosb_wage_basis = "flagged"
        st.save()

        p, _ = create_person(
            account=acc, first_name_ar="سعد", family_name_ar="القحطاني",
            gender="male", nationality_code="SA", id_type="national_id",
            id_number="1099887766", mobile="0509887766", force=True)
        emp, _, _ = create_employment(
            person=p, company=comp, employee_no="201",
            join_date=date(2019, 1, 1), iban=IBAN,
            salary_lines=[(comps["BASIC"], D("9000")),
                          (comps["HOUSING"], D("2250"))])
        yield {"account_id": r.account_id, "acc": acc, "comp": comp,
               "emp": emp, "person": p, "settings": st}


def _client(env, role_code, username="u", scope=Scope.COMPANY):
    u = User.objects.create_user(username=username, password="x")
    with account_scope(env["account_id"]):
        role = Role.objects.get(account_id=env["account_id"], code=role_code)
        m = AccountMembership.objects.create(
            user=u, account_id=env["account_id"], active_company=env["comp"])
        RoleAssignment.objects.create(membership=m, role=role,
                                      company=env["comp"], scope=scope.value)
    c = Client()
    c.force_login(u)
    return c


# ══════════ السجل ══════════

def test_all_reports_registered():
    load_reports()
    assert len(REGISTRY) >= 13


def test_every_report_declares_essentials():
    """كل تقرير يعلن مفتاحه وعنوانه ومجموعته وصلاحيته."""
    load_reports()
    for key, cls in REGISTRY.items():
        assert cls.key == key
        assert cls.title_ar, f"تقرير بلا عنوان: {key}"
        assert cls.group, f"تقرير بلا مجموعة: {key}"
        assert cls.permission, f"تقرير بلا صلاحية: {key}"


def test_catalog_grouped():
    load_reports()
    groups = {g["group"] for g in catalog()}
    assert {"financial", "attendance", "leaves", "employees"} <= groups


# ══════════ مخصصات نهاية الخدمة (ق-43) ══════════

@pytest.mark.django_db(transaction=True)
def test_eosb_provision_uses_full_award(env):
    """ق-43: المكافأة كاملة — تقدير للميزانية لا استحقاق فعلي."""
    with account_scope(env["account_id"]):
        res = get_report("eosb_provision")(
            company=env["comp"], as_of="2026-08-31").run()
        assert res.row_count == 1
        row = res.rows[0]
        # 7.67 سنة × 9000 — نصف شهر لأول خمس ثم شهر كامل
        assert D(row["eosb_provision"]) > D("40000")


@pytest.mark.django_db(transaction=True)
def test_eosb_provision_date_matters(env):
    """ق-43: تاريخ محدد لا فترة — الالتزام ينمو مع الزمن."""
    with account_scope(env["account_id"]):
        early = get_report("eosb_provision")(
            company=env["comp"], as_of="2022-01-01").run()
        late = get_report("eosb_provision")(
            company=env["comp"], as_of="2026-08-31").run()
        assert (D(late.totals["eosb_provision"])
                > D(early.totals["eosb_provision"]))


@pytest.mark.django_db(transaction=True)
def test_eosb_provision_includes_leave(env):
    """ق-43: المكافأة ورصيد الإجازات يُقاسان معًا."""
    with account_scope(env["account_id"]):
        res = get_report("eosb_provision")(
            company=env["comp"], as_of="2026-08-31").run()
        assert "leave_provision" in res.totals
        assert "total_provision" in res.totals


@pytest.mark.django_db(transaction=True)
def test_required_param_enforced(env):
    with account_scope(env["account_id"]):
        with pytest.raises(ReportError):
            get_report("eosb_provision")(company=env["comp"]).run()


# ══════════ تقارير أخرى ══════════

@pytest.mark.django_db(transaction=True)
def test_employees_report(env):
    with account_scope(env["account_id"]):
        res = get_report("employees")(
            company=env["comp"], status="active").run()
        assert res.row_count == 1
        assert res.rows[0]["employee_no"] == "201"


@pytest.mark.django_db(transaction=True)
def test_expiring_documents_report(env):
    with account_scope(env["account_id"]):
        add_document(employment=env["emp"], document_type="iqama",
                     expiry_date=date.today() + timedelta(days=20))
        res = get_report("expiring_documents")(
            company=env["comp"], within_days="60").run()
        assert res.row_count == 1
        assert res.rows[0]["severity"] in ("حرجة", "قريبة")


@pytest.mark.django_db(transaction=True)
def test_assets_report_totals_value(env):
    with account_scope(env["account_id"]):
        assign_asset(employment=env["emp"], name_ar="حاسب", value=D("4500"))
        assign_asset(employment=env["emp"], name_ar="هاتف", value=D("1500"))
        res = get_report("assets")(
            company=env["comp"], status="assigned").run()
        assert res.totals["value"] == "6000.00"


@pytest.mark.django_db(transaction=True)
def test_advances_report_empty_when_disabled(env):
    """ق-41: إطفاء النظام يُخفي السلف من التقارير."""
    with account_scope(env["account_id"]):
        env["settings"].advances_enabled = False
        env["settings"].save()
        res = get_report("advances")(company=env["comp"]).run()
        assert res.row_count == 0


# ══════════ التصدير ══════════

@pytest.mark.django_db(transaction=True)
def test_excel_export_is_valid(env):
    from openpyxl import load_workbook
    from apps.core.reports.excel import export_to_excel
    import io

    with account_scope(env["account_id"]):
        res = get_report("eosb_provision")(
            company=env["comp"], as_of="2026-08-31").run()
        data = export_to_excel(res, company_name="شركة التقارير")
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.sheet_view.rightToLeft is True
        assert ws.max_row > 5


@pytest.mark.django_db(transaction=True)
def test_excel_numbers_are_numeric(env):
    """المحاسب يجب أن يستطيع جمع الأعمدة."""
    from openpyxl import load_workbook
    from apps.core.reports.excel import export_to_excel
    import io

    with account_scope(env["account_id"]):
        res = get_report("eosb_provision")(
            company=env["comp"], as_of="2026-08-31").run()
        wb = load_workbook(io.BytesIO(
            export_to_excel(res, company_name="شركة")))
        ws = wb.active
        # عمود مخصص المكافأة (السابع) في أول صف بيانات
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=7).value
            if isinstance(v, (int, float)) and v > 1000:
                break
        else:
            pytest.fail("لا قيمة رقمية في عمود المخصص")


@pytest.mark.django_db(transaction=True)
def test_pdf_export_is_valid(env):
    from apps.core.reports.pdf import export_to_pdf

    with account_scope(env["account_id"]):
        res = get_report("eosb_provision")(
            company=env["comp"], as_of="2026-08-31").run()
        data = export_to_pdf(res, company_name="شركة التقارير")
        assert data[:5] == b"%PDF-"
        assert len(data) > 5000


def test_arabic_reshaping():
    """العربية في PDF تحتاج تشكيلًا وعكس اتجاه معًا."""
    from apps.core.reports.pdf import ar
    out = ar("تقرير")
    assert out != "تقرير"          # تغيّر فعلًا
    assert ar("12,345.67") == "12,345.67"   # الأرقام كما هي
    assert ar("") == ""


# ══════════ API ══════════

@pytest.mark.django_db(transaction=True)
def test_catalog_hides_unauthorized(env):
    """التقرير الذي لا يملك صلاحيته لا يظهر له أصلًا."""
    hr = _client(env, "hr_manager", "hr1")
    emp = _client(env, "employee", "emp1", Scope.OWN)

    hr_keys = {r["key"] for g in hr.get("/api/reports/").json()["groups"]
               for r in g["reports"]}
    emp_keys = {r["key"] for g in emp.get("/api/reports/").json()["groups"]
                for r in g["reports"]}
    assert "eosb_provision" in hr_keys
    assert len(emp_keys) < len(hr_keys)


@pytest.mark.django_db(transaction=True)
def test_run_report_json(env):
    c = _client(env, "hr_manager")
    d = c.get("/api/reports/eosb_provision/?as_of=2026-08-31").json()
    assert d["key"] == "eosb_provision"
    assert d["row_count"] == 1
    assert d["totals"]["total_provision"]


@pytest.mark.django_db(transaction=True)
def test_run_report_missing_param(env):
    c = _client(env, "hr_manager")
    r = c.get("/api/reports/eosb_provision/")
    assert r.status_code == 400
    assert r.json()["code"] == "missing_params"


@pytest.mark.django_db(transaction=True)
def test_unknown_report_404(env):
    c = _client(env, "hr_manager")
    r = c.get("/api/reports/made_up/")
    assert r.status_code == 404


@pytest.mark.django_db(transaction=True)
def test_download_xlsx(env):
    c = _client(env, "hr_manager")
    r = c.get("/api/reports/eosb_provision/?as_of=2026-08-31&export=xlsx")
    assert r.status_code == 200
    assert "attachment" in r["Content-Disposition"]
    assert ".xlsx" in r["Content-Disposition"]


@pytest.mark.django_db(transaction=True)
def test_download_pdf(env):
    c = _client(env, "hr_manager")
    r = c.get("/api/reports/eosb_provision/?as_of=2026-08-31&export=pdf")
    assert r.status_code == 200
    assert r["Content-Type"] == "application/pdf"


@pytest.mark.django_db(transaction=True)
def test_unsupported_format(env):
    c = _client(env, "hr_manager")
    r = c.get("/api/reports/eosb_provision/?as_of=2026-08-31&export=docx")
    assert r.status_code == 400


@pytest.mark.django_db(transaction=True)
def test_employee_cannot_run_payroll_report(env):
    emp = _client(env, "employee", "emp2", Scope.OWN)
    r = emp.get("/api/reports/eosb_provision/?as_of=2026-08-31")
    assert r.status_code == 403


@pytest.mark.django_db(transaction=True)
def test_reports_isolated_between_accounts(env):
    other = provision_account(slug="rpt-other", display_name_ar="آخر",
                              company_name_ar="أخرى", is_sandbox=True)
    with account_scope(other.account_id):
        from apps.employees.services.hiring import create_person as cp
        cp(account=Account.objects.get(id=other.account_id),
           first_name_ar="آخر", family_name_ar="شخص", gender="male",
           nationality_code="SA", id_type="national_id",
           id_number="1055443322", mobile="0505443322", force=True)

    c = _client(env, "hr_manager")
    d = c.get("/api/reports/employees/?status=active").json()
    assert d["row_count"] == 1
    assert all(r["employee_no"] == "201" for r in d["rows"])
