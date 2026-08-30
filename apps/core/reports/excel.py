"""
تصدير التقارير إلى Excel.

يعمل على أي تقرير يرث Report — إضافة تقرير جديد لا تمس هذا الملف.
"""
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SUB_FONT = Font(size=10, color="666666")
TOTAL_FILL = PatternFill("solid", fgColor="EEF3F6")
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FMT = "#,##0.00"
NUMBER_FMT = "#,##0.00"


def _coerce(value, kind):
    """يحوّل القيمة لنوعها الصحيح ليعالجها إكسل كرقم لا كنص."""
    if value in (None, ""):
        return ""
    if kind in ("money", "number"):
        try:
            return float(Decimal(str(value)))
        except (InvalidOperation, TypeError, ValueError):
            return value
    if kind == "date":
        if isinstance(value, (date, datetime)):
            return value
        try:
            return date.fromisoformat(str(value))
        except ValueError:
            return value
    return str(value)


def export_to_excel(result, company_name=""):
    """
    يبني ملف إكسل من نتيجة تقرير.

    الاتجاه من اليمين لليسار، والأرقام أرقامًا لا نصوصًا فيستطيع
    المحاسب أن يعالجها.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = result.title_ar[:31] or "تقرير"
    ws.sheet_view.rightToLeft = True

    ncols = max(1, len(result.columns))
    row = 1

    # ── الترويسة ──
    ws.cell(row=row, column=1, value=result.title_ar).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    row += 1

    if company_name:
        ws.cell(row=row, column=1, value=company_name).font = SUB_FONT
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ncols)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        row += 1

    if result.subtitle_ar:
        ws.cell(row=row, column=1, value=result.subtitle_ar).font = SUB_FONT
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ncols)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        row += 1

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(row=row, column=1,
            value=f"تاريخ الإصدار: {stamp}").font = SUB_FONT
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    row += 2

    # ── العناوين ──
    header_row = row
    for i, col in enumerate(result.columns, start=1):
        cell = ws.cell(row=row, column=i, value=col.label_ar)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = col.width
    ws.row_dimensions[row].height = 26
    row += 1

    # ── الصفوف ──
    for data in result.rows:
        for i, col in enumerate(result.columns, start=1):
            cell = ws.cell(row=row, column=i,
                           value=_coerce(data.get(col.key), col.kind))
            cell.border = BORDER
            if col.kind in ("money", "number"):
                cell.number_format = (MONEY_FMT if col.kind == "money"
                                      else NUMBER_FMT)
                cell.alignment = Alignment(horizontal="left")
            elif col.kind == "date":
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
        row += 1

    # ── الإجماليات ──
    if result.totals:
        for i, col in enumerate(result.columns, start=1):
            cell = ws.cell(row=row, column=i)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER
            if i == 1:
                cell.value = "الإجمالي"
                cell.alignment = Alignment(horizontal="right")
            elif col.key in result.totals:
                cell.value = float(result.totals[col.key])
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="left")
        row += 1

    # ── الملاحظات ──
    if result.notes:
        row += 1
        for note in result.notes:
            ws.cell(row=row, column=1, value=f"• {note}").font = SUB_FONT
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=ncols)
            ws.cell(row=row, column=1).alignment = Alignment(
                horizontal="right")
            row += 1

    # تجميد الترويسة وتفعيل الفلترة
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if result.rows:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(ncols)}"
            f"{header_row + len(result.rows)}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_filename(result):
    stamp = datetime.now().strftime("%Y%m%d")
    safe = result.key.replace("/", "_")
    return f"{safe}_{stamp}.xlsx"
